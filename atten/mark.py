import face_recognition
import cv2
import xlrd
import openpyxl
import datetime
import sys
import pickle
import os

cname=str(sys.argv[1])

userhome = os.path.expanduser('~')

face_enc = userhome + "\Documents\AttendanceManager\data" + "/" + cname + ".data"

xlfile = userhome + "\Documents\AttendanceManager\Records" + "/" + cname + ".xlsx"


s=datetime.datetime.now()
d=str(s.day)+"/"+str(s.month)+"/"+str(s.year)

wb = xlrd.open_workbook(xlfile)
sheet = wb.sheet_by_index(0)
sheet.cell_value(0, 0)

n=sheet.ncols+1

w = openpyxl.load_workbook(xlfile)
ws = w.active
ws.cell(1,n).value=d

known_face_names = []

for i in range(1,sheet.nrows):
    imname=sheet.cell_value(i,0)
    known_face_names.append(imname)


with open(face_enc,'rb') as filehandle:
    known_face_encodings = pickle.load(filehandle)

filehandle.close()

video_capture = cv2.VideoCapture(0)




# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

while True:
    # Grab a single frame of video
    ret, frame = video_capture.read()

    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]

    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            # See if the face is a match for the known face(s)
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            # If a match was found in known_face_encodings, just use the first one.
            if True in matches:
                first_match_index = matches.index(True)
                name = known_face_names[first_match_index]
                ws.cell(first_match_index+2,n).value='P'

            face_names.append(name)

    process_this_frame = not process_this_frame


    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        # Scale back up face locations since the frame we detected in was scaled to 1/4 size
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        # Draw a box around the face
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

        # Draw a label with a name below the face
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        font = cv2.FONT_HERSHEY_DUPLEX
        cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

    # Display the resulting image
    cv2.imshow('Video', frame)

    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
w.save(xlfile)
w.close()


 